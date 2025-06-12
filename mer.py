import cv2                                              # OpenCV for image processing
import pytesseract                                      # Tesseract OCR for text extraction
from ultralytics import YOLO                            # YOLO for object detection
import re                                               # Regex for text cleaning
import numpy as np                                      # NumPy for array operations
import os                                               # OS for file and folder operations
import pandas as pd                                     # Pandas for data manipulation
from tkinter import Tk, filedialog                      # Tkinter for file dialog GUI
from openpyxl import Workbook                           # openpyxl for Excel file creation
from openpyxl.utils.dataframe import dataframe_to_rows  # Convert DataFrame to Excel rows
from openpyxl.styles import Font                        # Excel font styling
from openpyxl.utils import get_column_letter            # Get Excel column letters
from openpyxl.worksheet.hyperlink import Hyperlink      # Add hyperlinks to Excel
from openpyxl.styles import Alignment                   # Excel cell alignment

CONFIDENCE_THRESHOLD = 0.3                              # Detection confidence threshold
logomod = "D:/ABC TOLL/LOGO177/weights/best.pt"         # Path to logo detection model
carcol = "D:/ABC TOLL/CARCOL/content/runs/detect/train2/weights/best.pt"  # Car color model
licol = "D:/ABC TOLL/LICOLI/classify/train2/weights/best.pt"              # Plate color model
output_folder = "D:/ABC TOLL/REPORT"                    # Output folder for reports
stickers_folder = os.path.join(output_folder, "stickers") # Folder for cropped stickers

def clean_text(text):
    return re.sub(r'[^a-zA-Z0-9]', '', text)            # Remove all non-alphanumeric chars from the text

def process_ocr_text(ocr_text):
    return clean_text(ocr_text.upper())                 # Convert text to uppercase and clean it

# Crop license plate, classify color, OCR text, and append info to results_dict
def crop_and_display_license_plate(image, license_plate_box):
    global results_dict                                 # Use the global results_dict variable
    x1, y1, x2, y2 = license_plate_box                  # Unpack bounding box coordinates
    license_plate_image = image[int(y1):int(y2), int(x1):int(x2)]  # Crop license plate from image
    model = YOLO(licol)                                 # Load license plate color classification model
    print("Performing license plate detection...")      # Debug print
    results = model.predict(license_plate_image, conf=0.5) # Predict license plate color
    top1_confidence = results[0].probs.top1conf         # Get top-1 confidence (not used further)
    top1_class_idx = results[0].probs.top1              # Get top-1 class index
    best_classification = results[0].names[top1_class_idx] # Get class name (color)
    tesseract_text = pytesseract.image_to_string(license_plate_image, config='--psm 7') # OCR on plate
    tesseract_text = clean_text(tesseract_text)         # Clean OCR text
    ocr_text = process_ocr_text(tesseract_text)         # Further process OCR text
    results_dict["objects_detected"].append({           # Append license plate info to results
        "class_name": "license-plate",
        "bounding_box": [x1, y1, x2, y2],
        "text": ocr_text,
        "color": best_classification
    })

# Crop car, classify color, and return info dict
def crop_car(image, car_box):
    x1, y1, x2, y2 = car_box                            # Unpack bounding box coordinates
    car_image = image[int(y1):int(y2), int(x1):int(x2)] # Crop car from image
    model = YOLO(carcol)                                # Load car color detection model
    results = model.predict(car_image, conf=0.5)        # Predict car color
    if results and len(results) > 0:                    # If results exist
        results = results[0]                            # Use first result
        class_idx = int(results.boxes.cls[0])           # Get class index of detected car color
        class_name = results.names[class_idx]           # Get class name (color)
        return {
            "class_name": "car",
            "bounding_box": [x1, y1, x2, y2],
            "color": class_name
        }
    return {
        "class_name": "car",
        "color": "unknown"
    }

# Crop logo, classify manufacturer, and return info dict
def crop_logo(image, logo_box):
    x1, y1, x2, y2 = logo_box                           # Unpack bounding box coordinates
    logo_image = image[int(y1):int(y2), int(x1):int(x2)]# Crop logo from image
    model = YOLO(logomod)                               # Load logo detection model
    results = model.predict(logo_image, conf=0.5)       # Predict logo
    top1_class_idx = int(results[0].probs.top1)         # Get top-1 class index
    best_classification = results[0].names[top1_class_idx] # Get class name (manufacturer)
    return {
        "class_name": "logo",
        "bounding_box": [x1, y1, x2, y2],
        "manufacturer": best_classification
    }

# Run detection on image, extract info, and return results_dict
def process_image(image, image_path):
    results_dict = {"objects_detected": [], "sticker_count": 0, "flag_count": 0} # Initialize results
    model = YOLO('D:/ABC TOLL/runs_archive/segment/train3/weights/best.pt')      # Load main detection model
    results = model(image, conf=CONFIDENCE_THRESHOLD)                            # Run detection
    if isinstance(results, list):                                                # If results is a list
        results = results[0]                                                     # Use first result
    names_dict = results.names                                                   # Get class names
    boxes = results.boxes.xyxy.tolist()                                          # Get bounding boxes
    class_labels = results.boxes.cls.tolist()                                    # Get class indices
    sticker_count = 0                                                            # Initialize sticker count
    flag_count = 0                                                               # Initialize flag count
    for i, box in enumerate(boxes):                                              # Iterate over detected objects
        class_idx = int(class_labels[i])                                         # Get class index
        class_name = names_dict[class_idx]                                       # Get class name
        if class_name == "sticker":                                              # If sticker detected
            sticker_count += 1
            crop_and_save(image, box, "sticker", results_dict, image_path)       # Crop and save sticker
        elif class_name == "flag":                                               # If flag detected
            flag_count += 1
            crop_and_save(image, box, "flag", results_dict, image_path)          # Crop and save flag
        if class_name == "license-plate":                                        # If license plate detected
            plate_info = crop_and_display_license_plate(image, box)              # Process license plate
            results_dict["objects_detected"].append(plate_info)                  # Append info
        if class_name == "car":                                                  # If car detected
            car_info = crop_car(image, box)                                      # Process car
            results_dict["objects_detected"].append(car_info)                    # Append info
        if class_name == "logo":                                                 # If logo detected
            logo_info = crop_logo(image, box)                                    # Process logo
            results_dict["objects_detected"].append(logo_info)                   # Append info
        # Save vehicle type if detected
        if class_name in ["car", "lorry", "bus", "auto", "ambulance"]:           # If vehicle type detected
            results_dict["vehicle_type"] = class_name                            # Save vehicle type
    results_dict["sticker_count"] = sticker_count                                # Update sticker count
    results_dict["flag_count"] = flag_count                                      # Update flag count
    # If no logo detected, add unknown
    if not any(item["class_name"] == "logo" for item in results_dict["objects_detected"]):
        results_dict["objects_detected"].append({
            "class_name": "logo",
            "manufacturer": "unknown"
        })
    # If no car detected, add unknown
    if not any(item["class_name"] == "car" for item in results_dict["objects_detected"]):
        results_dict["objects_detected"].append({
            "class_name": "car",
            "color": "unknown"
        })
    return results_dict

# Crop and save sticker/flag images, update results_dict with sticker paths
def crop_and_save(image, box, object_type, results_dict, image_path):
    x1, y1, x2, y2 = box                                # Unpack bounding box coordinates
    cropped_image = image[int(y1):int(y2), int(x1):int(x2)] # Crop object from image
    image_name = os.path.splitext(os.path.basename(image_path))[0] # Get image name without extension
    vehicle_folder = os.path.join(stickers_folder, image_name)     # Create folder for this vehicle
    os.makedirs(vehicle_folder, exist_ok=True)           # Ensure folder exists
    filename = f"{vehicle_folder}/{object_type}_{x1}_{y1}_{x2}_{y2}.jpg" # Build filename
    cv2.imwrite(filename, cropped_image)                 # Save cropped image
    if object_type == "sticker":                         # If object is a sticker
        if "stickers" not in results_dict:               # If stickers list not in results_dict
            results_dict["stickers"] = []                # Create stickers list
        results_dict["stickers"].append(filename)        # Add sticker path to results_dict

# GUI: Browse for folder, process all images, collect results, and save to Excel
def browse_and_process_images():
    Tk().withdraw()                                     # Hide main Tk window
    folder_selected = filedialog.askdirectory()         # Open folder dialog for user to select folder
    results_list = []                                   # List to store results for each image
    for filename in os.listdir(folder_selected):        # Loop through files in selected folder
        if filename.endswith((".jpg", ".jpeg", ".png")):# Only process image files
            image_path = os.path.join(folder_selected, filename) # Full path to image
            image = cv2.imread(image_path)              # Read image
            result = process_image(image, image_path)   # Process image and get results
            result["filename"] = filename               # Add filename to result
            result["filepath"] = image_path             # Add filepath to result
            results_list.append(result)                 # Add result to list
    save_results_to_excel(results_list, folder_selected)# Save all results to Excel

# Save results to Excel, add hyperlinks to images and sticker folders
def save_results_to_excel(results, folder_path):
    data = []                                           # List to hold row data for DataFrame
    sticker_folders_map = {}                            # Map filenames to sticker folder paths
    for result in results:                              # Loop through each result
        filename = result["filename"]                   # Get filename
        filepath = result["filepath"]                   # Get filepath
        objects_detected = result["objects_detected"]   # Get detected objects
        sticker_count = result["sticker_count"]         # Get sticker count
        flag_count = result["flag_count"]               # Get flag count
        vehicle_info = {                                # Dict to hold info for this vehicle
            "vehicle_image": filepath,
            "sticker_count": sticker_count,
            "flag_count": flag_count,
            "vehicle_type": result.get("vehicle_type", "unknown")
        }
        for obj in objects_detected:                    # Loop through detected objects
            if obj["class_name"] == "license-plate":    # If license plate detected
                vehicle_info.update({
                    "vehicle_number": obj["text"],
                    "license_plate_color": obj["color"]
                })
            elif obj["class_name"] == "car":            # If car detected
                vehicle_info["vehicle_color"] = obj["color"]
            elif obj["class_name"] == "logo":           # If logo detected
                vehicle_info["vehicle_logo"] = obj["manufacturer"]
        if "stickers" in result:                        # If stickers present
            sticker_folders_map[filename] = os.path.join(stickers_folder, os.path.splitext(filename)[0])
        data.append(vehicle_info)                       # Add vehicle info to data list
    df = pd.DataFrame(data)                             # Create DataFrame from data
    df["stickers_folder"] = ""                          # Add stickers_folder column
    output_file = os.path.join(folder_path, "report.xlsx") # Output Excel file path
    wb = Workbook()                                     # Create new Excel workbook
    ws = wb.active                                      # Get active worksheet
    for r in dataframe_to_rows(df, index=False, header=True): # Write DataFrame rows to worksheet
        ws.append(r)
    # Add hyperlinks for vehicle images and sticker folders
    for idx, row in df.iterrows():                      # Loop through DataFrame rows
        filepath = row["vehicle_image"]                 # Get image path
        filename = os.path.basename(filepath)           # Get image filename
        cell = ws.cell(row=idx + 2, column=1)           # Get cell for image
        cell.hyperlink = Hyperlink(ref=cell.coordinate, target=filepath) # Add hyperlink to image
        cell.style = "Hyperlink"                        # Set cell style
        cell.value = filename                           # Set cell value to filename
        sticker_folder = sticker_folders_map.get(filename, "") # Get sticker folder path
        if sticker_folder:                              # If sticker folder exists
            cell = ws.cell(row=idx + 2, column=len(df.columns)) # Get cell for sticker folder
            cell.hyperlink = Hyperlink(ref=cell.coordinate, target=sticker_folder) # Add hyperlink
            cell.style = "Hyperlink"                    # Set cell style
            cell.value = "Stickers"                     # Set cell value
    wb.save(output_file)                                # Save Excel file

browse_and_process_images()                             # Start the process when script runs
