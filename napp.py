import cv2                                              # OpenCV for image processing
import pytesseract                                      # Tesseract OCR for text extraction
from ultralytics import YOLO                            # YOLO for object detection
import re                                               # Regex for text cleaning
import numpy as np                                      # NumPy for array operations
import os                                               # OS for file and folder operations
import pandas as pd                                     # Pandas for data manipulation
from flask import Flask, request, jsonify               # Flask for web API
from openpyxl import Workbook                           # openpyxl for Excel file creation
from openpyxl.utils.dataframe import dataframe_to_rows  # Convert DataFrame to Excel rows
from openpyxl.styles import Font                        # Excel font styling
from openpyxl.utils import get_column_letter            # Get Excel column letters
from openpyxl.worksheet.hyperlink import Hyperlink      # Add hyperlinks to Excel
from openpyxl.styles import Alignment                   # Excel cell alignment

app = Flask(__name__)                                   # Create Flask app instance

# Confidence threshold for object detection
CONFIDENCE_THRESHOLD = 0.3                              # Set detection confidence threshold
logomod = "D:/ABC TOLL/LOGO177/weights/best.pt"         # Path to logo detection model
carcol = "D:/ABC TOLL/CARCOL/content/runs/detect/train2/weights/best.pt"  # Car color model
licol = "D:/ABC TOLL/LICOLI/classify/train2/weights/best.pt"              # Plate color model

# Function to clean OCR text
def clean_text(text):
    return re.sub(r'[^a-zA-Z0-9]', '', text)            # Remove all non-alphanumeric characters

# Function to process OCR text for license plate
def process_ocr_text(ocr_text):
    return clean_text(ocr_text.upper())                 # Convert text to uppercase and clean it

# Function to extract and display license plate information
def crop_and_display_license_plate(image, license_plate_box):
    x1, y1, x2, y2 = license_plate_box                  # Unpack bounding box coordinates
    license_plate_image = image[int(y1):int(y2), int(x1):int(x2)]  # Crop license plate from image
    model = YOLO(licol)                                 # Load license plate color classification model
    results = model.predict(license_plate_image, conf=0.5) # Predict license plate color
    top1_class_idx = int(results[0].probs.top1)         # Get top-1 class index
    best_classification = results[0].names[top1_class_idx] # Get class name (color)
    tesseract_text = pytesseract.image_to_string(license_plate_image, config='--psm 7') # OCR on plate
    ocr_text = process_ocr_text(tesseract_text)         # Clean and process OCR text

    return {                                            # Return license plate info as dict
        "class_name": "license-plate",
        "bounding_box": [x1, y1, x2, y2],
        "text": ocr_text,
        "color": best_classification
    }

# Function to extract and display car color information
def crop_car(image, car_box):
    x1, y1, x2, y2 = car_box                            # Unpack bounding box coordinates
    car_image = image[int(y1):int(y2), int(x1):int(x2)] # Crop car from image
    model = YOLO(carcol)                                # Load car color detection model
    results = model.predict(car_image, conf=0.5)        # Predict car color

    if results and len(results) > 0:                    # If results exist
        results = results[0]                            # Use first result
        class_idx = int(results.boxes.cls[0])           # Get class index of detected car color
        class_name = results.names[class_idx]           # Get class name (color)
        return {
            "class_name": "car",
            "bounding_box": [x1, y1, x2, y2],
            "color": class_name
        }
    return {                                            # If no car detected, return unknown
        "class_name": "car",
        "color": "unknown"
    }

# Function to extract and display logo information
def crop_logo(image, logo_box):
    x1, y1, x2, y2 = logo_box                           # Unpack bounding box coordinates
    logo_image = image[int(y1):int(y2), int(x1):int(x2)]# Crop logo from image
    model = YOLO(logomod)                               # Load logo detection model
    results = model.predict(logo_image, conf=0.5)       # Predict logo
    top1_class_idx = int(results[0].probs.top1)         # Get top-1 class index
    best_classification = results[0].names[top1_class_idx] # Get class name (manufacturer)

    return {                                            # Return logo info as dict
        "class_name": "logo",
        "bounding_box": [x1, y1, x2, y2],
        "manufacturer": best_classification
    }

# Process each image and return analysis result
def process_image(image, image_path, stickers_folder):
    results_dict = {"objects_detected": [], "sticker_count": 0, "flag_count": 0} # Initialize results dict
    model = YOLO('D:/ABC TOLL/runs_archive/segment/train3/weights/best.pt')      # Load main detection model
    results = model(image, conf=CONFIDENCE_THRESHOLD)                            # Run detection

    if isinstance(results, list):                                                # If results is a list
        results = results[0]                                                     # Use first result

    names_dict = results.names                                                   # Get class names
    boxes = results.boxes.xyxy.tolist()                                          # Get bounding boxes
    class_labels = results.boxes.cls.tolist()                                    # Get class indices

    sticker_count = 0                                                            # Initialize sticker count
    flag_count = 0                                                               # Initialize flag count

    for i, box in enumerate(boxes):                                              # Iterate over detected objects
        class_idx = int(class_labels[i])                                         # Get class index
        class_name = names_dict[class_idx]                                       # Get class name

        if class_name == "sticker":                                              # If sticker detected
            sticker_count += 1
            crop_and_save(image, box, "sticker", results_dict, image_path, stickers_folder) # Crop and save sticker
        elif class_name == "flag":                                               # If flag detected
            flag_count += 1
            crop_and_save(image, box, "flag", results_dict, image_path, stickers_folder)    # Crop and save flag

        if class_name == "license-plate":                                        # If license plate detected
            plate_info = crop_and_display_license_plate(image, box)              # Process license plate
            results_dict["objects_detected"].append(plate_info)                  # Append info

        if class_name == "car":                                                  # If car detected
            car_info = crop_car(image, box)                                      # Process car
            results_dict["objects_detected"].append(car_info)                    # Append info

        if class_name == "logo":                                                 # If logo detected
            logo_info = crop_logo(image, box)                                    # Process logo
            results_dict["objects_detected"].append(logo_info)                   # Append info

        # Add vehicle type if detected
        if class_name in ["car", "lorry", "bus", "auto", "ambulance"]:
            results_dict["vehicle_type"] = class_name

    results_dict["sticker_count"] = sticker_count                                # Update sticker count
    results_dict["flag_count"] = flag_count                                      # Update flag count

    # If no logo detected, add unknown
    if not any(item["class_name"] == "logo" for item in results_dict["objects_detected"]):
        results_dict["objects_detected"].append({
            "class_name": "logo",
            "manufacturer": "unknown"
        })

    # If no car detected, add unknown
    if not any(item["class_name"] == "car" for item in results_dict["objects_detected"]):
        results_dict["objects_detected"].append({
            "class_name": "car",
            "color": "unknown"
        })

    return results_dict

# Function to crop and save stickers and flags
def crop_and_save(image, box, object_type, results_dict, image_path, stickers_folder):
    x1, y1, x2, y2 = box                                # Unpack bounding box coordinates
    cropped_image = image[int(y1):int(y2), int(x1):int(x2)] # Crop object from image
    image_name = os.path.splitext(os.path.basename(image_path))[0] # Get image name without extension
    vehicle_folder = os.path.join(stickers_folder, image_name)     # Create folder for this vehicle
    os.makedirs(vehicle_folder, exist_ok=True)           # Ensure folder exists
    filename = f"{vehicle_folder}/{object_type}_{x1}_{y1}_{x2}_{y2}.jpg" # Build filename
    cv2.imwrite(filename, cropped_image)                 # Save cropped image
    if object_type == "sticker":                         # If object is a sticker
        if "stickers" not in results_dict:               # If stickers list not in results_dict
            results_dict["stickers"] = []                # Create stickers list
        results_dict["stickers"].append(filename)        # Add sticker path to results_dict

# Function to save results to Excel
def save_results_to_excel(results, folder_path, stickers_folder):
    data = []                                           # List to hold row data for DataFrame
    sticker_folders_map = {}                            # Map filenames to sticker folder paths
    for result in results:                              # Loop through each result
        filename = result["filename"]                   # Get filename
        filepath = result["filepath"]                   # Get filepath
        objects_detected = result["objects_detected"]   # Get detected objects
        sticker_count = result["sticker_count"]         # Get sticker count
        flag_count = result["flag_count"]               # Get flag count
        vehicle_info = {                                # Dict to hold info for this vehicle
            "vehicle_image": filepath,
            "sticker_count": sticker_count,
            "flag_count": flag_count,
            "vehicle_type": result.get("vehicle_type", "unknown")
        }
        for obj in objects_detected:                    # Loop through detected objects
            if obj["class_name"] == "license-plate":    # If license plate detected
                vehicle_info.update({
                    "vehicle_number": obj["text"],
                    "license_plate_color": obj["color"]
                })
            elif obj["class_name"] == "car":            # If car detected
                vehicle_info["vehicle_color"] = obj["color"]
            elif obj["class_name"] == "logo":           # If logo detected
                vehicle_info["vehicle_logo"] = obj["manufacturer"]
        if "stickers" in result:                        # If stickers present
            sticker_folders_map[filename] = os.path.join(stickers_folder, os.path.splitext(filename)[0])
        data.append(vehicle_info)                       # Add vehicle info to data list
    
    df = pd.DataFrame(data)                             # Create DataFrame from data
    df["stickers_folder"] = ""                          # Add stickers_folder column

    output_file = os.path.join(folder_path, "report.xlsx") # Output Excel file path
    
    # Create a workbook and sheet
    wb = Workbook()                                     # Create new Excel workbook
    ws = wb.active                                      # Get active worksheet
    
    # Write DataFrame to Excel sheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Add hyperlinks and stickers to the Excel sheet
    for idx, row in df.iterrows():
        filepath = row["vehicle_image"]                 # Get image path
        filename = os.path.basename(filepath)           # Get image filename
        
        # Hyperlink for vehicle image
        cell = ws.cell(row=idx + 2, column=1)           # Get cell for image
        cell.hyperlink = Hyperlink(ref=cell.coordinate, target=filepath) # Add hyperlink to image
        cell.style = "Hyperlink"                        # Set cell style
        cell.value = filename                           # Set cell value to filename

        # Stickers folder column
        sticker_folder = sticker_folders_map.get(filename, "")
        if sticker_folder:
            cell = ws.cell(row=idx + 2, column=len(df.columns))
            cell.hyperlink = Hyperlink(ref=cell.coordinate, target=sticker_folder)
            cell.style = "Hyperlink"
            cell.value = "Stickers"
    
    # Save the workbook
    wb.save(output_file)

@app.route('/process-images', methods=['POST'])          # Define Flask route for processing images
def process_images():
    folder_path = request.json.get('folder_path')        # Get folder path from POST request JSON
    if not folder_path or not os.path.exists(folder_path):# Validate folder path
        return jsonify({"error": "Invalid folder path"}), 400
    
    stickers_folder = os.path.join(folder_path, "stickers") # Path for saving stickers
    os.makedirs(stickers_folder, exist_ok=True)          # Ensure stickers folder exists

    results_list = []                                    # List to store results for each image
    for filename in os.listdir(folder_path):             # Loop through files in folder
        if filename.endswith((".jpg", ".jpeg", ".png")): # Only process image files
            image_path = os.path.join(folder_path, filename) # Full path to image
            image = cv2.imread(image_path)               # Read image
            result = process_image(image, image_path, stickers_folder) # Analyze image
            result["filename"] = filename                # Add filename to result
            result["filepath"] = image_path              # Add filepath to result
            results_list.append(result)                  # Add result to list
    
    save_results_to_excel(results_list, folder_path, stickers_folder) # Save all results to Excel
    return jsonify({"message": "Processing completed successfully"}), 200 # Return success

if __name__ == '__main__':                              # If script is run directly
    app.run(debug=True)                                 # Start Flask app in debug mode
