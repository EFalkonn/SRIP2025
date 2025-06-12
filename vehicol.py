import cv2                                               # Import OpenCV for image processing tasks
import pytesseract                                       # Import Tesseract OCR for text extraction from images
from ultralytics import YOLO                             # Import YOLO from ultralytics for object detection
import re                                                # Import regex module for text processing and cleaning
import numpy as np                                       # Import NumPy for numerical operations, especially array manipulation
import os                                                # Import OS module for interacting with the operating system (file/folder operations)
import pandas as pd                                      # Import Pandas for data manipulation and analysis, particularly DataFrames
from flask import Flask, request, jsonify                # Import Flask, request, and jsonify for creating a web API
from openpyxl import Workbook                            # Import Workbook from openpyxl to create Excel files
from openpyxl.utils.dataframe import dataframe_to_rows   # Import utility to convert Pandas DataFrame to rows for Excel
from openpyxl.styles import Font                         # Import Font for styling text in Excel cells
from openpyxl.utils import get_column_letter             # Import utility to get Excel column letters (e.g., A, B, C)
from openpyxl.worksheet.hyperlink import Hyperlink       # Import Hyperlink to add hyperlinks in Excel cells
from openpyxl.styles import Alignment                    # Import Alignment for setting cell alignment in Excel
from datetime import datetime                            # Import datetime to work with dates and times, used for timestamping

app = Flask(__name__)                                    # Create an instance of the Flask application

# Confidence threshold for object detection
CONFIDENCE_THRESHOLD = 0.3                               # Set the minimum confidence score for an object detection to be considered valid
logomod = "D:/ABC TOLL/LOGO177/weights/best.pt"          # Path to the pre-trained YOLO model for logo detection
carcol = "D:/ABC TOLL/carcolll/best.pt"                  # Path to the pre-trained YOLO model for car color detection
licol = "D:/ABC TOLL/LICOLI/classify/train2/weights/best.pt"              # Path to the pre-trained YOLO model for license plate color classification

# Function to clean OCR text
def clean_text(text):
    return re.sub(r'[^a-zA-Z0-9]', '', text)            # Remove all characters that are not alphanumeric

# Function to process OCR text for license plate
def process_ocr_text(ocr_text):
    return clean_text(ocr_text.upper())                 # Convert text to uppercase and then clean it

# Function to extract and display license plate information
def crop_and_display_license_plate(image, license_plate_box):
    x1, y1, x2, y2 = license_plate_box                  # Unpack the bounding box coordinates (top-left x, top-left y, bottom-right x, bottom-right y)
    license_plate_image = image[int(y1):int(y2), int(x1):int(x2)]  # Crop the license plate region from the image
    model = YOLO(licol)                                 # Load the license plate color classification model
    results = model.predict(license_plate_image, conf=0.5) # Perform prediction on the cropped license plate image

    if results and len(results) > 0 and results[0].probs is not None and results[0].probs.top1 is not None: # Check if results are valid and top1 probability exists
        top1_class_idx = int(results[0].probs.top1)     # Get the index of the class with the highest probability
        best_classification = results[0].names[top1_class_idx] # Get the name of the best classified color
        tesseract_text = pytesseract.image_to_string(license_plate_image, config='--psm 7') # Perform OCR on the license plate image
        ocr_text = process_ocr_text(tesseract_text)     # Process the extracted OCR text
        return {                                        # Return a dictionary with license plate information
            "class_name": "license-plate",
            "bounding_box": [x1, y1, x2, y2],
            "text": ocr_text,
            "color": best_classification
        }

    return {                                            # Return default "unknown" values if detection or OCR fails
        "class_name": "license-plate",
        "text": "unknown",
        "color": "unknown"
    }

# Function to extract and display car color information
def crop_car(image, car_box):
    x1, y1, x2, y2 = car_box                            # Unpack the bounding box coordinates
    car_image = image[int(y1):int(y2), int(x1):int(x2)] # Crop the car region from the image
    model = YOLO(carcol)                                # Load the car color detection model
    results = model.predict(car_image, conf=0.5)        # Perform prediction on the cropped car image

    if results and len(results) > 0 and results[0].boxes.cls is not None and len(results[0].boxes.cls) > 0: # Check if results are valid and class predictions exist
        results = results[0]                            # Use the first set of results (assuming one dominant car color)
        class_idx = int(results[0].boxes.cls[0])        # Get the class index of the detected car color
        class_name = results.names[class_idx]           # Get the name of the detected color
        return {                                        # Return a dictionary with car color information
            "class_name": "car",                        # Set class name to "car" for identification
            "bounding_box": [x1, y1, x2, y2],           # Store the bounding box of the car
            "color": class_name                         # Store the detected color
        }
    return {                                            # Return default "unknown" color if detection fails
        "class_name": "car",
        "color": "unknown"
    }

# Function to extract and display logo information
def crop_logo(image, logo_box):
    x1, y1, x2, y2 = logo_box                           # Unpack the bounding box coordinates
    logo_image = image[int(y1):int(y2), int(x1):int(x2)]# Crop the logo region from the image
    model = YOLO(logomod)                               # Load the logo detection model
    results = model.predict(logo_image, conf=0.5)       # Perform prediction on the cropped logo image

    if results and len(results) > 0 and results[0].probs is not None and results[0].probs.top1 is not None: # Check if results are valid and top1 probability exists
        top1_class_idx = int(results[0].probs.top1)     # Get the index of the class with the highest probability
        best_classification = results[0].names[top1_class_idx] # Get the name of the best classified logo/manufacturer
        return {                                        # Return a dictionary with logo information
            "class_name": "logo",                       # Set class name to "logo"
            "bounding_box": [x1, y1, x2, y2],           # Store the bounding box of the logo
            "manufacturer": best_classification         # Store the detected manufacturer
        }

    return {                                            # Return default "unknown" manufacturer if detection fails
        "class_name": "logo",
        "manufacturer": "unknown"
    }

# Process each image and return analysis result
def process_image(image, image_path):
    results_dict = {"objects_detected": [], "flag_count": 0} # Initialize a dictionary to store detection results, including a flag counter
    # Load the main YOLO model for general object detection (vehicles, plates, flags, etc.)
    model = YOLO('D:/ABC TOLL/runs_archive/segment/train3/weights/best.pt')
    results = model(image, conf=CONFIDENCE_THRESHOLD)    # Perform object detection on the input image using the set confidence threshold
    
    if isinstance(results, list):                        # Ensure results are in the expected format (a single Results object)
        results = results[0]                             # Take the first element if results is a list

    names_dict = results.names                           # Get a dictionary mapping class indices to class names
    boxes = results.boxes.xyxy.tolist()                  # Get bounding box coordinates for detected objects
    class_labels = results.boxes.cls.tolist()            # Get class labels (indices) for detected objects

    flag_count = 0                                       # Initialize flag counter

    for i, box in enumerate(boxes):                      # Iterate through each detected object's bounding box
        class_idx = int(class_labels[i])                 # Get the class index for the current object
        class_name = names_dict[class_idx]               # Get the class name for the current object

        if class_name == "flag":                         # If the detected object is a flag
            flag_count += 1                              # Increment flag count

        if class_name == "license-plate":                # If the detected object is a license plate
            plate_info = crop_and_display_license_plate(image, box) # Process the license plate (crop, OCR, color)
            results_dict["objects_detected"].append(plate_info)     # Add license plate info to the results dictionary

        if class_name in ["car", "lorry", "bus", "auto"]:            # If the detected object is a car or other specified vehicle type, process it for color
            car_info = crop_car(image, box)              # Process the car (crop, get color)
            results_dict["objects_detected"].append(car_info)       # Add car info (color) to the results dictionary

        if class_name == "logo":                         # If the detected object is a logo
            logo_info = crop_logo(image, box)            # Process the logo (crop, get manufacturer)
            results_dict["objects_detected"].append(logo_info)      # Add logo info to the results dictionary

        # Add vehicle type based on the main model's detection
        if class_name in ["car", "lorry", "bus", "auto", "ambulance"]: # Check if the class name is a known vehicle type
            results_dict["vehicle_type"] = class_name    # Store the detected vehicle type in the results dictionary

    results_dict["flag_count"] = flag_count              # Store the final flag count in the results dictionary

    # If no logo was specifically detected and processed, add a default "unknown" entry for logo
    if not any(item["class_name"] == "logo" for item in results_dict["objects_detected"]):
        results_dict["objects_detected"].append({
            "class_name": "logo",
            "manufacturer": "unknown"
        })

    # If no car color was specifically detected and processed, add a default "unknown" entry for car color
    # This ensures the 'car' entry exists for consistent Excel reporting, even if only vehicle type was found.
    if not any(item["class_name"] == "car" and "color" in item for item in results_dict["objects_detected"]):
        # Check if a 'car' entry already exists (e.g. from vehicle_type) and update it, or add a new one.
        car_entry_exists = False
        for item in results_dict["objects_detected"]:
            if item["class_name"] == "car":
                item["color"] = item.get("color", "unknown") # Ensure color field exists
                car_entry_exists = True
                break
        if not car_entry_exists:
             results_dict["objects_detected"].append({
                "class_name": "car", # This entry is primarily for color, vehicle type is separate
                "color": "unknown"
            })


    return results_dict                                  # Return the dictionary containing all processed information

# Function to save results to Excel
def save_results_to_excel(results, folder_path):
    data = []                                           # Initialize a list to store data for the DataFrame
    for result in results:                              # Iterate through each processed image's result
        filename = result["filename"]                   # Get the original filename of the image
        filepath = result["filepath"]                   # Get the full filepath of the image
        objects_detected = result["objects_detected"]   # Get the list of detected objects and their details
        flag_count = result["flag_count"]               # Get the count of detected flags
        vehicle_info = {                                # Create a dictionary to store consolidated vehicle information for the Excel row
            "vehicle_image": filepath,                  # Store the path to the vehicle image
            "flag_count": flag_count,                   # Store the number of flags detected
            "vehicle_type": result.get("vehicle_type", "unknown") # Get vehicle type, default to "unknown" if not found
        }
        # Iterate through each detected object's details to populate vehicle_info
        for obj in objects_detected:
            if obj["class_name"] == "license-plate":    # If it's a license plate
                vehicle_info.update({                   # Update vehicle_info with plate number and color
                    "vehicle_number": obj.get("text", "unknown"),
                    "license_plate_color": obj.get("color", "unknown")
                })
            elif obj["class_name"] == "car":            # If it's a car (this entry is mainly for color)
                vehicle_info["vehicle_color"] = obj.get("color", "unknown") # Add car color
            elif obj["class_name"] == "logo":           # If it's a logo
                vehicle_info["vehicle_logo"] = obj.get("manufacturer", "unknown") # Add logo manufacturer
        data.append(vehicle_info)                       # Add the consolidated vehicle information for this image to the data list
    
    df = pd.DataFrame(data)                             # Create a Pandas DataFrame from the collected data

    # Generate the timestamp for the report filename
    timestamp = datetime.now().strftime("%d-%m-%Y_%H.%M.%S") # Format: DD-MM-YYYY_HH.MM.SS
    output_file = os.path.join(folder_path, f"report_{timestamp}.xlsx") # Define the output Excel filename with timestamp
    
    # Create a new Excel Workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicle Analysis Report"                # Set the title of the worksheet
    
    # Write DataFrame to Excel sheet, including headers
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_idx, value in enumerate(row):
            ws.cell(row=r_idx + 1, column=c_idx + 1, value=value) # Write cell value

    # Add hyperlinks for vehicle images in the Excel sheet
    if not df.empty:                                    # Proceed only if DataFrame has data
        header_offset = 1                               # Headers are in the first row
        # Find the column index for 'vehicle_image' (1-based for openpyxl)
        try:
            vehicle_image_col_idx = df.columns.get_loc("vehicle_image") + 1
        except KeyError:
            print("Warning: 'vehicle_image' column not found in DataFrame. Hyperlinks for images will not be created.")
            vehicle_image_col_idx = None

        if vehicle_image_col_idx:
            for idx, row_data in df.iterrows():         # Iterate through each row of the DataFrame (idx is 0-based)
                excel_row_num = idx + 1 + header_offset # Calculate Excel row number (1-based, plus header)
                
                filepath = row_data["vehicle_image"]    # Get the vehicle image filepath from the DataFrame
                img_cell = ws.cell(row=excel_row_num, column=vehicle_image_col_idx) # Get the cell for the vehicle image link
                img_cell.value = os.path.basename(filepath) # Set cell text to the filename
                img_cell.hyperlink = Hyperlink(ref=img_cell.coordinate, target=filepath) # Create hyperlink to the image file
                img_cell.style = "Hyperlink"            # Apply hyperlink style
    
    # Auto-adjust column widths for better readability
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column) # Get the column letter (e.g., A, B)
        for cell in col:
            try:                                        # Handle empty cells or non-string values
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = (max_length + 2) if max_length > 0 else 12 # Set a minimum width if column is empty
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_file)                                # Save the Excel workbook

@app.route('/process-images', methods=['POST'])          # Define a Flask route that accepts POST requests at '/process-images'
def process_images_route():                             # Renamed function to avoid conflict with any potential 'process_images' variable
    if not request.is_json:                             # Check if the request content type is JSON
        return jsonify({"error": "Request must be JSON"}), 400 # Return error if not JSON
    
    data = request.get_json()                           # Get JSON data from the request
    folder_path = data.get('folder_path')               # Extract 'folder_path' from JSON data

    if not folder_path or not os.path.isdir(folder_path):# Validate if folder_path is provided and is a valid directory
        return jsonify({"error": "Invalid or missing folder_path"}), 400 # Return error for invalid path
    
    results_list = []                                    # Initialize a list to store results for all processed images
    # Iterate through each file in the specified folder_path
    for filename in os.listdir(folder_path):
        # Check if the file is an image (jpg, jpeg, png) - case-insensitive
        if filename.lower().endswith((".jpg", ".jpeg", ".png")):
            image_path = os.path.join(folder_path, filename) # Construct the full path to the image
            try:
                image = cv2.imread(image_path)           # Read the image using OpenCV
                if image is None:                        # Check if image reading failed
                    print(f"Warning: Could not read image {filename}. Skipping.")
                    continue                             # Skip to the next file
                
                # Process the image to detect objects and extract information
                result = process_image(image, image_path)
                result["filename"] = filename            # Add the original filename to the result dictionary
                result["filepath"] = image_path          # Add the full filepath to the result dictionary
                results_list.append(result)              # Add the result to the list of results
            except Exception as e:
                print(f"Error processing image {filename}: {e}") # Log any error during image processing
    
    if not results_list:                                 # Check if no images were processed or found
        return jsonify({"message": "No images found or processed in the specified folder."}), 200

    try:
        # Save the collected results to an Excel file
        save_results_to_excel(results_list, folder_path)
        return jsonify({"message": "Processing completed successfully. Report generated."}), 200 # Return success message
    except Exception as e:
        print(f"Error saving results to Excel: {e}")    # Log any error during Excel saving
        return jsonify({"error": f"Failed to save Excel report: {e}"}), 500 # Return server error

if __name__ == '__main__':                              # Check if the script is being run directly (not imported as a module)
    app.run(debug=True, host='0.0.0.0', port=5000)      # Run the Flask development server, accessible on the network, with debugging enabled
