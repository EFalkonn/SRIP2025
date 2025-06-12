import cv2                                               # Import OpenCV for image processing tasks
import pytesseract                                       # Import Tesseract OCR for text extraction from images
from ultralytics import YOLO                             # Import YOLO from ultralytics for object detection
import re                                                # Import regex module for text processing and cleaning
import numpy as np                                       # Import NumPy for numerical operations, especially array manipulation
import os                                                # Import OS module for interacting with the operating system (file/folder operations)
import pandas as pd                                      # Import Pandas for data manipulation and analysis, particularly DataFrames
from flask import Flask, request, jsonify                # Import Flask, request, and jsonify for creating a web API
from openpyxl import Workbook                            # Import Workbook from openpyxl to create Excel files
from openpyxl.utils.dataframe import dataframe_to_rows   # Import utility to convert Pandas DataFrame to rows for Excel
from openpyxl.styles import Font                         # Import Font for styling text in Excel cells
from openpyxl.utils import get_column_letter             # Import utility to get Excel column letters (e.g., A, B, C)
from openpyxl.worksheet.hyperlink import Hyperlink       # Import Hyperlink to add hyperlinks in Excel cells
from openpyxl.styles import Alignment                    # Import Alignment for setting cell alignment in Excel
from datetime import datetime                            # Import datetime to work with dates and times, used for timestamping

app = Flask(__name__)                                    # Create an instance of the Flask application

# Confidence threshold for object detection
CONFIDENCE_THRESHOLD = 0.3                               # Set the minimum confidence score for an object detection to be considered valid
logomod = "D:/ABC TOLL/LOGO177/weights/best.pt"          # Path to the pre-trained YOLO model for logo detection
carcol = "D:/ABC TOLL/CARCOL/content/runs/detect/train2/weights/best.pt"  # Path to the pre-trained YOLO model for car color detection
licol = "D:/ABC TOLL/LICOLI/classify/train2/weights/best.pt"              # Path to the pre-trained YOLO model for license plate color classification

# Function to clean OCR text
def clean_text(text):
    return re.sub(r'[^a-zA-Z0-9]', '', text)            # Remove all characters that are not alphanumeric

# Function to process OCR text for license plate
def process_ocr_text(ocr_text):
    return clean_text(ocr_text.upper())                 # Convert text to uppercase and then clean it

# Function to extract and display license plate information
def crop_and_display_license_plate(image, license_plate_box):
    x1, y1, x2, y2 = license_plate_box                  # Unpack the bounding box coordinates (top-left x, top-left y, bottom-right x, bottom-right y)
    license_plate_image = image[int(y1):int(y2), int(x1):int(x2)]  # Crop the license plate region from the image
    model = YOLO(licol)                                 # Load the license plate color classification model
    results = model.predict(license_plate_image, conf=0.5) # Perform prediction on the cropped license plate image

    if results and len(results) > 0 and results[0].probs is not None and results[0].probs.top1 is not None: # Check if results are valid and top1 probability exists
        top1_class_idx = int(results[0].probs.top1)     # Get the index of the class with the highest probability
        best_classification = results[0].names[top1_class_idx] # Get the name of the best classified color
        tesseract_text = pytesseract.image_to_string(license_plate_image, config='--psm 7') # Perform OCR on the license plate image
        ocr_text = process_ocr_text(tesseract_text)     # Process the extracted OCR text
        return {                                        # Return a dictionary with license plate information
            "class_name": "license-plate",
            "bounding_box": [x1, y1, x2, y2],
            "text": ocr_text,
            "color": best_classification
        }

    return {                                            # Return default "unknown" values if detection or OCR fails
        "class_name": "license-plate",
        "text": "unknown",
        "color": "unknown"
    }

# Function to extract and display car color information
def crop_car(image, car_box):
    x1, y1, x2, y2 = car_box                            # Unpack the bounding box coordinates
    car_image = image[int(y1):int(y2), int(x1):int(x2)] # Crop the car region from the image
    model = YOLO(carcol)                                # Load the car color detection model
    results = model.predict(car_image, conf=0.5)        # Perform prediction on the cropped car image

    if results and len(results) > 0 and results[0].boxes.cls is not None and len(results[0].boxes.cls) > 0: # Check if results are valid and class predictions exist
        results = results[0]                            # Use the first set of results (assuming one dominant car color)
        class_idx = int(results.boxes.cls[0])           # Get the class index of the detected car color
        class_name = results.names[class_idx]           # Get the name of the detected color
        return {                                        # Return a dictionary with car color information
            "class_name": "car",
            "bounding_box": [x1, y1, x2, y2],
            "color": class_name
        }
    return {                                            # Return default "unknown" color if detection fails
        "class_name": "car",
        "color": "unknown"
    }

# Function to extract and display logo information
def crop_logo(image, logo_box):
    x1, y1, x2, y2 = logo_box                           # Unpack the bounding box coordinates
    logo_image = image[int(y1):int(y2), int(x1):int(x2)]# Crop the logo region from the image
    model = YOLO(logomod)                               # Load the logo detection model
    results = model.predict(logo_image, conf=0.5)       # Perform prediction on the cropped logo image

    if results and len(results) > 0 and results[0].probs is not None and results[0].probs.top1 is not None: # Check if results are valid and top1 probability exists
        top1_class_idx = int(results[0].probs.top1)     # Get the index of the class with the highest probability
        best_classification = results[0].names[top1_class_idx] # Get the name of the best classified logo/manufacturer
        return {                                        # Return a dictionary with logo information
            "class_name": "logo",
            "bounding_box": [x1, y1, x2, y2],
            "manufacturer": best_classification
        }

    return {                                            # Return default "unknown" manufacturer if detection fails
        "class_name": "logo",
        "manufacturer": "unknown"
    }

# Process each image and return analysis result
def process_image(image, image_path, stickers_folder):
    results_dict = {"objects_detected": [], "sticker_count": 0, "flag_count": 0} # Initialize a dictionary to store detection results
    # Load the main YOLO model for general object detection (vehicles, plates, stickers, flags)
    model = YOLO('D:/ABC TOLL/runs_archive/segment/train3/weights/best.pt')
    results = model(image, conf=CONFIDENCE_THRESHOLD)    # Perform object detection on the input image

    if isinstance(results, list):                        # Ensure results are in the expected format (a single Results object)
        results = results[0]                             # Take the first element if results is a list

    names_dict = results.names                           # Get a dictionary mapping class indices to class names
    boxes = results.boxes.xyxy.tolist()                  # Get bounding box coordinates for detected objects
    class_labels = results.boxes.cls.tolist()            # Get class labels (indices) for detected objects

    sticker_count = 0                                    # Initialize sticker counter
    flag_count = 0                                       # Initialize flag counter

    for i, box in enumerate(boxes):                      # Iterate through each detected object's bounding box
        class_idx = int(class_labels[i])                 # Get the class index for the current object
        class_name = names_dict[class_idx]               # Get the class name for the current object

        if class_name == "sticker":                      # If the detected object is a sticker
            sticker_count += 1                           # Increment sticker count
            crop_and_save(image, box, "sticker", results_dict, image_path, stickers_folder) # Crop and save the sticker image
        elif class_name == "flag":                       # If the detected object is a flag
            flag_count += 1                              # Increment flag count
            # crop_and_save(image, box, "flag", results_dict, image_path, stickers_folder) # Optionally crop and save flag images

        if class_name == "license-plate":                # If the detected object is a license plate
            plate_info = crop_and_display_license_plate(image, box) # Process the license plate
            results_dict["objects_detected"].append(plate_info)     # Add license plate info to results

        if class_name == "car":                          # If the detected object is a car
            car_info = crop_car(image, box)              # Process the car (for color)
            results_dict["objects_detected"].append(car_info)       # Add car info to results

        if class_name == "logo":                         # If the detected object is a logo
            logo_info = crop_logo(image, box)            # Process the logo
            results_dict["objects_detected"].append(logo_info)      # Add logo info to results

        # Add vehicle type if detected by the main model
        if class_name in ["car", "lorry", "bus", "auto", "ambulance"]: # Check if the class name is a known vehicle type
            results_dict["vehicle_type"] = class_name    # Store the detected vehicle type

    results_dict["sticker_count"] = sticker_count        # Store the final sticker count
    results_dict["flag_count"] = flag_count              # Store the final flag count

    # If no logo was specifically detected and processed, add a default "unknown" entry
    if not any(item["class_name"] == "logo" for item in results_dict["objects_detected"]):
        results_dict["objects_detected"].append({
            "class_name": "logo",
            "manufacturer": "unknown"
        })

    # If no car was specifically detected and processed for color, add a default "unknown" entry
    if not any(item["class_name"] == "car" for item in results_dict["objects_detected"]):
        results_dict["objects_detected"].append({
            "class_name": "car",
            "color": "unknown"
        })

    return results_dict                                  # Return the dictionary containing all processed information

# Function to crop and save stickers and flags
def crop_and_save(image, box, object_type, results_dict, image_path, stickers_folder):
    x1, y1, x2, y2 = box                                # Unpack bounding box coordinates
    cropped_image = image[int(y1):int(y2), int(x1):int(x2)] # Crop the object from the main image
    image_name = os.path.splitext(os.path.basename(image_path))[0] # Get the base name of the original image (without extension)
    vehicle_folder = os.path.join(stickers_folder, image_name)     # Create a subfolder path for this vehicle's stickers
    os.makedirs(vehicle_folder, exist_ok=True)           # Create the subfolder if it doesn't exist
    # Create a unique filename for the cropped image based on type and coordinates
    filename = f"{vehicle_folder}/{object_type}_{int(x1)}_{int(y1)}_{int(x2)}_{int(y2)}.jpg"
    cv2.imwrite(filename, cropped_image)                 # Save the cropped image
    if object_type == "sticker":                         # If the cropped object is a sticker
        if "stickers" not in results_dict:               # Check if the 'stickers' key exists in results_dict
            results_dict["stickers"] = []                # If not, initialize it as an empty list
        results_dict["stickers"].append(filename)        # Add the path of the saved sticker image to the list

# Function to save results to Excel
def save_results_to_excel(results, folder_path, stickers_folder):
    data = []                                           # Initialize a list to store data for the DataFrame
    sticker_folders_map = {}                            # Initialize a dictionary to map image filenames to their sticker folders
    for result in results:                              # Iterate through each processed image's result
        filename = result["filename"]                   # Get the original filename of the image
        filepath = result["filepath"]                   # Get the full filepath of the image
        objects_detected = result["objects_detected"]   # Get the list of detected objects and their details
        sticker_count = result["sticker_count"]         # Get the count of detected stickers
        flag_count = result["flag_count"]               # Get the count of detected flags
        vehicle_info = {                                # Create a dictionary to store consolidated vehicle information
            "vehicle_image": filepath,
            "sticker_count": sticker_count,
            "flag_count": flag_count,
            "vehicle_type": result.get("vehicle_type", "unknown") # Get vehicle type, default to "unknown"
        }
        for obj in objects_detected:                    # Iterate through each detected object's details
            if obj["class_name"] == "license-plate":    # If it's a license plate
                vehicle_info.update({                   # Update vehicle_info with plate number and color
                    "vehicle_number": obj.get("text", "unknown"),
                    "license_plate_color": obj.get("color", "unknown")
                })
            elif obj["class_name"] == "car":            # If it's a car (for color information)
                vehicle_info["vehicle_color"] = obj.get("color", "unknown") # Add car color
            elif obj["class_name"] == "logo":           # If it's a logo
                vehicle_info["vehicle_logo"] = obj.get("manufacturer", "unknown") # Add logo manufacturer
        if "stickers" in result and result["stickers"]: # If there are saved sticker images for this vehicle
            # Map the original image filename to the folder containing its cropped stickers
            sticker_folders_map[filename] = os.path.join(stickers_folder, os.path.splitext(filename)[0])
        data.append(vehicle_info)                       # Add the consolidated vehicle information to the data list
    
    df = pd.DataFrame(data)                             # Create a Pandas DataFrame from the collected data
    if not df.empty:                                    # Check if the DataFrame is not empty
         df["stickers_folder_link"] = ""                # Add a new column for sticker folder links, initialize as empty
    else: # Handle empty DataFrame case if necessary, e.g. by not creating "stickers_folder_link" or logging
        print("No data to save to Excel.")
        return


    # Generate the timestamp for the report filename
    timestamp = datetime.now().strftime("%d-%m-%Y_%H.%M.%S") # Format: DD-MM-YYYY_HH.MM.SS
    output_file = os.path.join(folder_path, f"report_{timestamp}.xlsx") # Define the output Excel filename with timestamp
    
    wb = Workbook()                                     # Create a new Excel Workbook
    ws = wb.active                                      # Get the active worksheet
    ws.title = "Vehicle Analysis Report"                # Set the title of the worksheet
    
    # Write DataFrame to Excel sheet, including headers
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_idx, value in enumerate(row):
            ws.cell(row=r_idx + 1, column=c_idx + 1, value=value) # Write cell value

    # Add hyperlinks for vehicle images and sticker folders
    if not df.empty: # Proceed only if DataFrame has data
        header_offset = 1 # Assuming headers are in the first row
        vehicle_image_col_idx = df.columns.get_loc("vehicle_image") + 1 # Get column index for vehicle_image
        stickers_folder_link_col_idx = df.columns.get_loc("stickers_folder_link") + 1 # Get column index for stickers_folder_link

        for idx, row_data in df.iterrows():             # Iterate through each row of the DataFrame (idx is 0-based)
            excel_row_num = idx + 1 + header_offset     # Calculate Excel row number (1-based, plus header)
            
            # Hyperlink for vehicle image
            filepath = row_data["vehicle_image"]        # Get the vehicle image filepath
            img_cell = ws.cell(row=excel_row_num, column=vehicle_image_col_idx) # Get the cell for the vehicle image link
            img_cell.value = os.path.basename(filepath) # Set cell text to the filename
            img_cell.hyperlink = Hyperlink(ref=img_cell.coordinate, target=filepath) # Create hyperlink to the image file
            img_cell.style = "Hyperlink"                # Apply hyperlink style

            # Hyperlink for stickers folder
            original_filename = os.path.basename(filepath) # Get the original image filename
            sticker_folder_path = sticker_folders_map.get(original_filename) # Get the path to the stickers folder
            if sticker_folder_path:                     # If a sticker folder exists for this image
                sf_cell = ws.cell(row=excel_row_num, column=stickers_folder_link_col_idx) # Get cell for sticker folder link
                sf_cell.value = "View Stickers"         # Set cell text
                sf_cell.hyperlink = Hyperlink(ref=sf_cell.coordinate, target=sticker_folder_path) # Create hyperlink
                sf_cell.style = "Hyperlink"             # Apply hyperlink style
    
    # Auto-adjust column widths for better readability
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column letter
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_file)                                # Save the Excel workbook

@app.route('/process-images', methods=['POST'])          # Define a Flask route that accepts POST requests
def process_images_route():
    if not request.is_json:                             # Check if the request content type is JSON
        return jsonify({"error": "Request must be JSON"}), 400 # Return error if not JSON
    
    data = request.get_json()                           # Get JSON data from the request
    folder_path = data.get('folder_path')               # Extract 'folder_path' from JSON data

    if not folder_path or not os.path.isdir(folder_path):# Validate if folder_path is provided and is a valid directory
        return jsonify({"error": "Invalid or missing folder_path"}), 400 # Return error for invalid path
    
    # Define the path for storing cropped stickers, relative to the input folder_path
    stickers_folder = os.path.join(folder_path, "processed_stickers_output")
    os.makedirs(stickers_folder, exist_ok=True)          # Create the stickers folder if it doesn't exist

    results_list = []                                    # Initialize a list to store results for all processed images
    # Iterate through each file in the specified folder_path
    for filename in os.listdir(folder_path):
        # Check if the file is an image (jpg, jpeg, png)
        if filename.lower().endswith((".jpg", ".jpeg", ".png")):
            image_path = os.path.join(folder_path, filename) # Construct the full path to the image
            try:
                image = cv2.imread(image_path)           # Read the image using OpenCV
                if image is None:                        # Check if image reading failed
                    print(f"Warning: Could not read image {filename}. Skipping.")
                    continue                             # Skip to the next file
                
                # Process the image to detect objects and extract information
                result = process_image(image, image_path, stickers_folder)
                result["filename"] = filename            # Add the original filename to the result dictionary
                result["filepath"] = image_path          # Add the full filepath to the result dictionary
                results_list.append(result)              # Add the result to the list of results
            except Exception as e:
                print(f"Error processing image {filename}: {e}") # Log any error during image processing
                # Optionally, add error information to results_list or handle differently
    
    if not results_list:                                 # Check if no images were processed or found
        return jsonify({"message": "No images found or processed in the specified folder."}), 200

    try:
        # Save the collected results to an Excel file
        save_results_to_excel(results_list, folder_path, stickers_folder)
        return jsonify({"message": "Processing completed successfully. Report generated."}), 200 # Return success message
    except Exception as e:
        print(f"Error saving results to Excel: {e}")    # Log any error during Excel saving
        return jsonify({"error": f"Failed to save Excel report: {e}"}), 500 # Return server error

if __name__ == '__main__':                              # Check if the script is being run directly
    app.run(debug=True, host='0.0.0.0', port=5000)      # Run the Flask development server (accessible on network)
